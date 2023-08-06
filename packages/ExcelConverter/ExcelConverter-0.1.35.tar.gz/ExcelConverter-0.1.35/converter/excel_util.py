import collections
import string
import numpy as np
import logging
from decimal import Decimal
from collections import Iterable
from openpyxl.styles.numbers import *

ERR_NUM = '#NUM!'
ERR_NULL = '#NULL!'
ERR_NA = '#N/A'
ERR_VALUE = '#VALUE!'
ERR_REF = '#REF!'
XLFN = '_xlfn.'
ERR_NAME = '#NAME?'
ERR_DIV = '#DIV/0!'

ERROR_VALUES = [ERR_DIV, ERR_NA, ERR_NAME, ERR_NULL, ERR_NUM, ERR_REF, ERR_VALUE]

LOG_FINE = 8

logger = logging.getLogger(__name__)


"""
Remove any trailing zeros as fractional part
Example: 5.500 -> 5.5
Ref: https://stackoverflow.com/questions/11227620/drop-trailing-zeros-from-decimal
"""
def normalize_fraction(d):
    if not isinstance(d, Decimal):
        d = Decimal(d)
    normalized = d.normalize()
    sign, digit, exponent = normalized.as_tuple()
    return normalized if exponent <= 0 else normalized.quantize(1)
"""
It will convert the list of rows to a dictionary such that the keys correspond 
to the first column in the list of row and the values to the column indicated through the 
argument col index.
"""
def to_dictionary(row_list, col_index_num):
    return {row[0]: row[col_index_num] for row in row_list}

def create_address(rng, sheet=None):
    if is_collection(rng):
        address = ''
        for rg in rng:
            if len(address) > 0:
                address += ';'
            address += create_address_single(rg, sheet)
        return address
    else:
        return create_address_single(rng, sheet)

def create_address_single(rng, sheet=None):
    address = rng.replace('$', '')
    if sheet is not None and address.find('!') < 0:
        address = sheet + '!' + address
    return address

def safe_str(value):
    return str(value) if value is not None else ''

def get_range(spreadsheet, cell_address):
    return spreadsheet.cellmap[cell_address].address()

def evaluate_spreadsheet(spreadsheet, sheet, col_nr, row_nr):
    return spreadsheet.evaluate(sheet + '!' + col_nr + str(row_nr))

def to_values(spreadsheet, rng, skip_none=True, is_valid_function=None):
    result = []
    def process(col_nr, row_nr, value):
        if not skip_none or value is not None and is_valid_function(value):
            result.append(value)
        return True
    iterate_cells(spreadsheet, rng, process)
    return result

# Change item from Uppercase to Lowercase
def lower_case_item(item):
    if isinstance(item, Iterable):
        item = [str(entry).lower() for entry in item]
        return item
    else:
        return str(item).lower()

"""
rng can be a single range string or a tuple of ranges.
"""
def iterate_cells(spreadsheet, rng, process_fn):
    if isinstance(rng, list):
        for rg in rng:
            iterate_cells_single(spreadsheet, rg, process_fn)
    else:
        iterate_cells_single(spreadsheet, rng, process_fn)

"""
process_fn will receive the absolute col and row, both in numbers. If returned, false, iteration will stop.
"""
def iterate_cells_single(spreadsheet, rng, process_fn):
    sheet, start, end = split_range(rng)
    sh, start_col, start_row = split_address(start)
    sh, end_col, end_row = split_address(end)
    start_col_nr, end_col_nr = col2num(start_col), col2num(end_col)
    start_row_nr, end_row_nr = int(start_row), int(end_row)
    for row_nr in range(start_row_nr, end_row_nr + 1):
        for col_nr in range(start_col_nr, end_col_nr + 1):
            try:
                if not process_fn(col_nr, row_nr, evaluate_spreadsheet(spreadsheet, sheet, num2col(col_nr), row_nr)):
                    return
            except:
                raise Exception("Error iterating: col_nr: [%s], row_nr: [%s], address: [%s]" % (col_nr, row_nr, rng))

def rankdata(a, method='average'):
    """
    Assign ranks to data, dealing with ties appropriately.
    Ranks begin at 1.  The `method` argument controls how ranks are assigned
    to equal values.  See [1]_ for further discussion of ranking methods.
    Parameters
    ----------
    a : array_like
        The array of values to be ranked.  The array is first flattened.
    method : str, optional
        The method used to assign ranks to tied elements.
        The options are 'average', 'min', 'max', 'dense' and 'ordinal'.
        'average':
            The average of the ranks that would have been assigned to
            all the tied values is assigned to each value.
        'min':
            The minimum of the ranks that would have been assigned to all
            the tied values is assigned to each value.  (This is also
            referred to as "competition" ranking.)
        'max':
            The maximum of the ranks that would have been assigned to all
            the tied values is assigned to each value.
        'dense':
            Like 'min', but the rank of the next highest element is assigned
            the rank immediately after those assigned to the tied elements.
        'ordinal':
            All values are given a distinct rank, corresponding to the order
            that the values occur in `a`.
        The default is 'average'.
    Returns
    -------
    ranks : ndarray
         An array of length equal to the size of `a`, containing rank
         scores.
    References
    ----------
    .. [1] "Ranking", http://en.wikipedia.org/wiki/Ranking
    .. [2] https://github.com/scipy/scipy/blob/master/scipy/stats/stats.py
    Examples
    --------
    """
    if method not in ('average', 'min', 'max', 'dense', 'ordinal'):
        raise ValueError('unknown method "{0}"'.format(method))

    arr = np.ravel(np.asarray(a))
    algo = 'mergesort' if method == 'ordinal' else 'quicksort'
    sorter = np.argsort(arr, kind=algo)

    inv = np.empty(sorter.size, dtype=np.intp)
    inv[sorter] = np.arange(sorter.size, dtype=np.intp)

    if method == 'ordinal':
        return inv + 1

    arr = arr[sorter]
    obs = np.r_[True, arr[1:] != arr[:-1]]
    dense = obs.cumsum()[inv]

    if method == 'dense':
        return dense

    # cumulative counts of each unique value
    count = np.r_[np.nonzero(obs)[0], len(obs)]

    if method == 'max':
        return count[dense]

    if method == 'min':
        return count[dense - 1] + 1

    # average method
    return .5 * (count[dense] + count[dense - 1] + 1)

def is_eval_allowed(self, cell):
    return not isinstance(cell.compiled_expression, str)
    # isinstance(cell.compiled_expression, str) and re.match('[A-Za-z]+', cell.compiled_expression)

def is_compile_allowed(self, obj):
    return obj.formula is not None

    # This doesn't work in case the value is a
    # string that is the result of a formula like a vlookup, as such that the value
    # will be copied to the compiled expression and used, such that updated will not occur.
    # return not isinstance(obj.value, str)
    #
    # isinstance(self.value, str) and re.match('[A-Za-z]+', self.value):

def create_string_from_cell_value(value):
    # If given value is a String, checks if value contains multiple lines.
    # If value does contain multiple lines, value is split into lines and
    # rebuild to a single String with \\n appended between every line.
    # If value is not a String, just the quotes are added if not already present
    # and the result is returned.
    if not isinstance(value, str):
        raise TypeError('Value is not a String')
    lines = value.splitlines()
    if len(lines) > 1:
        result = '\\n'.join(lines)
    else:
        result = value
    if not result.startswith('"'):
        result = '"' + result
    if not result.endswith('"'):
        result = result + '"'
    return result

def get_zeros(cell):
    if hasattr(cell, 'number_format') and cell.number_format is not None and '.' in cell.number_format:
        # We split the number_format on the dot so that we get a list like so: ['0', '00']
        list_of_zeros = list(cell.number_format.split('.'))
        # Retrieve length of the second index and return it
        length_of_zeros = len(list_of_zeros[1])
        return int(length_of_zeros)
    else: return None


def clean_formula(formula):
    return remove_xlfn(formula)

def remove_xlfn(formula):
    return formula.replace(XLFN, '', 1) if isinstance(formula, str) else formula

def correct_range(excel, rng):
    return [correct_range_single(excel, rg) for rg in rng] if is_collection(rng) else correct_range_single(excel, rng)

def correct_range_single(excel, rng, sheet=None):
    # Ensure the correct sheet is used
    sheet = create_sheet_plus(rng, sheet)
    if sheet is not None:
        excel.set_sheet(sheet)

    # TODONEdNew: retrieve max rows/cols from cache.
    return correct_range_max(rng, excel.max_columns(), excel.max_rows())

def correct_range_max(rng, max_col, max_row):
    rng = rng.replace('$', '')
    sheet, start, end = split_range(rng)
    if not start and not end:
        return rng

    sheet, start, end = split_range(rng)
    sh, start_col, start_row = split_address(start)
    sh, end_col, end_row = split_address(end)

    start_col_nr = col2num(start_col)
    end_col_nr = col2num(end_col)
    if start_col_nr > end_col_nr:
        raise ValueError("Incorrect range specified: [%s], max_col: [%s], max_row: [%s]" % (rng, max_col, max_row))
    mx_col_start = min(start_col_nr, max_col)
    mx_col_end = min(end_col_nr, max_col)

    if start_row is not None:
        start_row_nr = int(start_row)
        mx_row_start = min(start_row_nr, max_row)
    else:
        mx_row_start = 1

    if end_row is not None:
        end_row_nr = int(end_row)
        mx_row_end = min(end_row_nr, max_row)
    else:
        mx_row_end = max_row

    if mx_col_start < start_col_nr or mx_col_end < end_col_nr or start_row is None or mx_row_start < start_row_nr\
            or end_row is None or mx_row_end < end_row_nr:
        # We are requesting more cell and available in the excel, let's correct it
        result = num2col(mx_col_start) if start_col_nr > mx_col_start else str(start_col)
        result += str(mx_row_start) + ':'
        result += num2col(mx_col_end) if end_col_nr > mx_col_end  else str(end_col)
        result += str(mx_row_end)

        #        result =  mx_col_start + str(mx_row_start) + ':' + num2col(mx_col_end) + str(mx_row_end)
        return sheet + '!' + result if sheet else result

    return rng

class CellRange(object):
    def __init__(self, addresses, rng, value, sheet=None):
        self.__addresses = addresses
        self.__address = rng.replace('$', '')
        sh, start, end = split_range(self.__address)

        if not sh and not sheet:
            raise Exception("Must pass in a sheet")

        # make sure the address is always prefixed with the range
        if sh is not None:
            sheet = sh
        else:
            self.__address = sheet + "!" + self.__address

        sh, start_col, start_row = split_address(start)
        sh, end_col, end_row = split_address(end)

        # # don't allow messing with these params

        self.__nrows = int(end_row) - int(start_row)
        self.__ncols = col2num(end_col) - col2num(start_col)

        self.__sheet = sheet
        self.value = value

        # We assume that a new cell range isn't dirty as it's value will be set
        self.__dirty = False
        self.__eval_inprogress = False
        self.__reset_inprogress = False

    def __str__(self):
        return self.__address

    def address(self):
        return self.__address

    """
    Indicates the exact addresses contained in the ranges
    """
    def addresses(self):
        return self.__addresses

    """
    Used to indicate the cell range is dirty, which means re-evaluation is required
    """
    @property
    def dirty(self):
        return self.__dirty

    @dirty.setter
    def dirty(self, value):
        self.__dirty = value

    """
    Used to indicate the cell is being evaluated
    """
    @property
    def eval_inprogress(self):
        return self.__eval_inprogress

    @eval_inprogress.setter
    def eval_inprogress(self, value):
        self.__eval_inprogress = value

    """
    Used to indicate the cell is being reset
    """
    @property
    def reset_inprogress(self):
        return self.__reset_inprogress

    @reset_inprogress.setter
    def reset_inprogress(self, value):
        self.__reset_inprogress = value

    @property
    def nrows(self):
        return self.__nrows

    @property
    def ncols(self):
        return self.__ncols

    # Is required/used when adding a node to the graph
    @property
    def sheet(self):
        return self.__sheet


class Cell(object):
    ctr = 0
    # __dirty = False

    @classmethod
    def next_id(cls):
        cls.ctr += 1
        return cls.ctr


    def __init__(self, address, sheet, formula, value, number_format=None):
        super(Cell, self).__init__()

        # remove $'s
        address = address.replace('$', '')

        sh, c, r = split_address(address)

        # both are empty
        if not sheet and not sh:
            raise Exception("Sheet name may not be empty for cell address %s" % address)
        # both exist but disagree
        elif sh and sheet and sh != sheet:
            raise Exception("Sheet name mismatch for cell address %s: %s vs %s" % (address, sheet, sh))
        elif not sh and sheet:
            sh = sheet
        else:
            pass

        # we assume a cell's location can never change
        # __ Means that it is unchangeable
        self.__sheet = str(sheet)
        self.__sheet = sh
        self.__col = c
        self.__row = int(r) if r else None
        self.__col_idx = col2num(c)

        # Checks if formula is None, if so set formula to None

        if formula is not None:
            self.__formula = formula
            self.value = value
        elif formula is None:
            self.__formula = None
            self.value = value
        elif isinstance(self.value, str):
            self.value = value

        # We assume that a new cell isn't dirty as it's value is set above
        self.__dirty = False
        #  self.__dirty = self.value is None
        self.__eval_inprogress = False
        self.__reset_inprogress = False
        self.number_format = number_format
        self.python_expression = None
        self._compiled_expression = None

        # every cell has a unique id
        self.__id = Cell.next_id()

    @property
    def number_fmt(self):
        return self.number_format

    @number_fmt.setter
    def number_fmt(self, value):
        self.number_format = value

    """
    Used to indicate the cell is dirty, which means re-evaluation is required
    """
    @property
    def dirty(self):
        return self.__dirty

    @dirty.setter
    def dirty(self, value):
        self.__dirty = value

    """
    Used to indicate the cell is being evaluated
    """
    @property
    def eval_inprogress(self):
        return self.__eval_inprogress

    @eval_inprogress.setter
    def eval_inprogress(self, value):
        self.__eval_inprogress = value

    """
    Used to indicate the cell is being reset
    """
    @property
    def reset_inprogress(self):
        return self.__reset_inprogress

    @reset_inprogress.setter
    def reset_inprogress(self, value):
        self.__reset_inprogress = value

    @property
    def sheet(self):
        return self.__sheet

    @property
    def row(self):
        return self.__row

    @property
    def col(self):
        return self.__col

    @property
    def formula(self):
        return self.__formula

    @property
    def id(self):
        return self.__id

    @property
    def compiled_expression(self):
        return self._compiled_expression

    # code objects are not serializable
    def __getstate__(self):
        d = dict(self.__dict__)
        f = '_compiled_expression'
        if f in d:
            del d[f]
        return d

    def __setstate__(self, d):
        self.__dict__.update(d)
        self.compile()

    def clean_name(self):
        return self.address().replace('!', '_').replace(' ', '_')

    def address(self, absolute=True):
        if absolute is True:
            return "%s!%s%s" % (self.__sheet, self.__col, self.__row)
        else:
            return "%s%s" % (self.__col, self.__row)

    def address_parts(self):
        return self.__sheet, self.__col, self.__row, self.__col_idx

    def compile(self):
        if not self.python_expression:
            return

        # if we are a constant string, surround by quotes
        if (isinstance(self.value, str) or isinstance(self.value, int)) \
                and self.formula and self.python_expression.startswith('"'):
            self.python_expression = '"' + self.python_expression + '"'

        try:
            if is_compile_allowed(self, self):
                self._compiled_expression = compile(self.python_expression, '<string>', 'eval')
            else:
                self._compiled_expression = self.value
        except Exception as e:
            raise Exception(
                "Failed to compile cell %s with expression %s: %s" % (self.address(), self.python_expression, e))

    def __str__(self):
        if self.formula is not None:
            return "%s%s" % (self.address(), self.formula)
        else:
            return "%s=%s" % (self.address(), self.value)

    @staticmethod
    def inc_col_address(address, inc):
        sh, col, row = split_address(address)
        return "%s!%s%s" % (sh, num2col(col2num(col) + inc), row)

    @staticmethod
    def inc_row_address(address, inc):
        sh, col, row = split_address(address)
        return "%s!%s%s" % (sh, col, row + inc)

    @staticmethod
    def resolve_cell(excel, address, sheet=None):
        r = excel.get_range(address)
        f = r.Formula if r.Formula.startswith('=') else None
        v = r.value
        nf = r.number_format
        sh, c, r = split_address(address)

        # use the sheet specified in the cell, else the passed sheet
        if sh:
            sheet = sh

        c = Cell(address, sheet, value=v, formula=f, number_format=nf)
        return c

    @staticmethod
    def make_cells_multi(excel: object, ranges: object) -> object:
        result = []
        for sheet, rng in ranges.items():
            # starting points
            cursheet = sheet if sheet else excel.get_active_sheet()
            excel.set_sheet(cursheet)
            rng = correct_range_single(excel, rng,
                                       sheet)  # Ensure the correct sheet is active before correcting the range

            cells, nrows, ncols, rng = Cell.make_cells(excel, rng, cursheet)
            result.extend(list(flatten(cells)))
        return result

    @staticmethod
    def make_cells(excel: object, rng: object, sheet=None, parent_ast=None, cellmap=None) -> object:
        """
        :rtype: object
        """
        cells = []

        def convert_range(rng, sheet=None, parent_ast=None, cellmap=None):
            cells = []

            # use the sheet specified in the range, else the passed sheet
            sh, start, end = split_range(rng)
            if sh is not None:
                sheet = sh

            if logger.isEnabledFor(logging.DEBUG):
                logger.debug("Creating cells for range: [%s]", rng)

            ads, numrows, numcols, rgs = resolve_range(rng, sheet=sheet, parent_ast=parent_ast)
            address = create_address(rgs)
            if cellmap is not None and address in cellmap:
                return None, numrows, numcols, address

            if is_collection(rgs):
                rng = rgs
                if logger.isEnabledFor(logging.DEBUG):
                    logger.debug("Created ranges for range: [%s], ranges: [%s]", rng, rgs)

                fs, vs, nf = [], [], []
                for rn in rgs:
                    r = excel.get_range(rn)
                    fm = r.Formula
                    if is_collection(fm):
                        fs.append(list(flatten(fm)))
                    else:
                        fs.append([fm])

                    nm = r.number_format
                    if is_collection(nf):
                        nf.append(list(flatten(nm)))
                    else:
                        nf.append([nm])

                    va = r.value
                    if is_collection(va):
                        vs.append(list(flatten(va)))
                    else:
                        vs.append([va])

                if len(fs) != len(ads):
                    fs = list(zip(*fs))
                    vs = list(zip(*vs))
                    nf = list(zip(*nf))
            else:
                r = excel.get_range(rng)
                fs = r.Formula
                vs = r.value
                nf = r.number_format  # nf: NumberFormat
                # It must be a 2-dim array (row and column)
                if numrows == 1:
                    ads = [ads]
                elif numcols == 1:
                    ads = [[x] for x in ads]

            # In case the value is a single integer, we must make it a string else the zip will throw an error indicating
            # it's not iteratable.
            # In case vs is None we have to wrap it in an array as well, else the zip function below will fail as None
            # isn't iteratable
            if vs is None or isinstance(vs, (int, float, str)):
                vs = [[str(vs)]]

            # We put the formula in a list, as a single string might result in problems.
            # Encountered problem was: ads=[['AV5']], fs='=AV178' that resulted below in the zip in fs='='
            # Apparently unzipping (zip(*x)) will use the "=" as some kind of separator.
            if fs is None or isinstance(fs, str):
                fs = [[fs]]

            if nf is None or isinstance(nf, str):
                nf = [[nf]]

            try:
                # We can't merge ads, fs and vs as vs can be less that can result in cells being ignored
                # Example: last column contains only formulas such that vs returns rows with 1 column less then a fs row
                for ad_row_nr in range(len(ads)):
                    row = []
                    ad_row = ads[ad_row_nr]
                    for ad_cell_nr in range(len(ad_row)):
                        a = ad_row[ad_cell_nr]
                        if is_collection(a):
                            raise Exception("Collection not allowed: [%s], range: [%s], ranges: [%s], formulas: [%s], "
                                            "values: [%s]" % (a, rng, rgs, fs, vs))

                        formula = clean_formula(fs[ad_row_nr][ad_cell_nr])
                        formula = formula if formula and formula.startswith('=') else None

                        value = vs[ad_row_nr]
                        nr_format = nf[ad_row_nr]

                        # A value can be empty, example: the last column contains only formulas
                        value = value[ad_cell_nr] if ad_cell_nr < len(value) else None
                        nr_format = nr_format[ad_cell_nr] if ad_cell_nr < len(nr_format) else None
                        cl = Cell(a, sheet, value=value, formula=formula, number_format=nr_format)

                        # We expect a formula that is at least 2 character long, the first character should be the "="
                        if formula is not None and len(formula) < 2:
                            raise Exception(
                                "Wrong formula encountered: [%s], a: [%s], v: [%s], ads: [%s], fs: [%s], "
                                "vs: [%s], range: [%s], cell: [%s]" % (formula, a, value, ads, fs, vs, rng, cl))
                        row.append(cl)

                        if logger.isEnabledFor(LOG_FINE):
                            logger.log(LOG_FINE, "Created cell for range: [%s!%s], cell: [%s]", sheet, rng, cl)

                    if len(row) > 0:
                        cells.append(row)

                # return as vector
                if len(cells) > 0 and (numrows == 1 or numcols == 1):
                    if numrows == 1:
                        cells = cells[0]
                    else:
                        cls = []
                        for x in cells:
                            # We ignore empty cells, like cells that result in an excel #REF error.
                            if x is not None:
                                cls.append(x[0] if isinstance(x, list) else x)

                                # Removed: as it creates an error for empty/removed (#REF excel error) cells
                                # cells = [x[0] for x in cells]
                        cells = cls

                return cells, numrows, numcols, rng
            except Exception as ex:
                raise Exception(
                    "Error creating row cells for range [%s], sheet: [%s], ads: [%s], num_rows: [%s], num_cols: [%s], formula: [%s],"
                    " values: [%s], cells: [%s], number_format: [%s]" % (rng, sheet, ads, numrows, numcols, fs, vs, cells, nr_format)) from ex

        def create_range_exception(rng, cell, sheet):
            return create_exception(rng, cell, sheet, "Error converting")

        def create_resolve_exception(rng, cell, sheet):
            return create_exception(rng, cell, sheet, "Error resolving")

        def create_exception(rng, cell, sheet, msg):
            return Exception(msg + ", range: [%s], cell: [%s], sheet: [%s], excel: [%s]"
                             % (rng, cell, sheet, excel))

        if isinstance(rng, list):  # if a list of cells
            cells = []
            for cell in rng:
                if is_range(cell):
                    try:
                        cs_in_range, nr, nc, rng = convert_range(cell, sheet, cellmap)
                        cells.append(cs_in_range)
                    except Exception as ex:
                        raise create_range_exception(rng, cell, sheet) from ex
                else:
                    try:
                        c = Cell.resolve_cell(excel, cell, sheet=sheet)
                        cells.append(c)
                    except Exception as ex:
                        raise create_resolve_exception(rng, cell, sheet) from ex

            cells = list(flatten(cells))
            return cells, -1, -1, None  # numrows, numcols and range are irrelevant here, so we return nr=nc=-1
        else:
            if is_range(rng):
                try:
                    # We return the range as it will be specific in case the shorthand range notitation was used,
                    # shorthand: (A:B), and specific: (A88:B99)
                    cells, numrows, numcols, rng = convert_range(rng, sheet, parent_ast, cellmap)
                except Exception as ex:
                    raise create_range_exception(rng, None, sheet) from ex
            else:
                try:
                    c = Cell.resolve_cell(excel, rng, sheet=sheet)
                    cells.append(c)
                    numrows = 1
                    numcols = 1
                except Exception as ex:
                    raise create_resolve_exception(rng, None, sheet) from ex

        return cells, numrows, numcols, rng

def is_range(address):
    return address.find(':') > 0

def split_range(rng):
    try:
        sh, start, end = None, None, None
        if rng.find('!') > 0:
            sh, r = rng.split("!")
            if rng.find(':') > 0:
                start, end = r.split(':')
        else:
            sh = None
            if rng.find(':') > 0:
                start, end = rng.split(':')

        return sh, start, end
    except Exception as ex:
        raise Exception("Error spliting range: [%s]" % (rng)) from ex

def create_sheet(rng):
    if rng.find('!') > 0:
        sh, r = rng.split("!")
        return sh
    return None

"""
It includes the '!' symbol at the end.
In case the value of sheet is specified, it must match the sheet value in the specified range.
We use the specified sheet, in case it's present and not present in the specified range.
"""
def create_sheet_plus(rng, sheet=None, add_close=False):
    sh, start, end = split_range(rng)
    if sh and sheet is not None:
        if sh != sheet:
            sheet = sh
            # sh = sheet
            # raise Exception("Mismatched sheets %s and %s" % (sh, sheet))
    elif sh is not None and not sheet:
        sheet = sh

    if sheet and add_close:
        sheet += '!'
    return sheet

def split_address(address):
    sheet = None
    if address.find('!') > 0:
        sheet, address = address.split('!')

    # ignore case
    address = address.upper()

    # regular <col><row> format    
    if re.match('^[A-Z $]+[\d $]+$', address):
        col, row = [_f for _f in re.split('([A-Z $]+)', address) if _f]
    # shorthand range notition (A:H) instead of (A1:H100) (al rows in A till all rows in H)
    elif re.match('^[A-Z $]+$', address):
        col = address
        row = None
    # R<row>C<col> format
    elif re.match('^R\d+C\d+$', address):
        row, col = address.split('C')
        row = row[1:]
    # R[<row>]C[<col>] format
    elif re.match('^R\[\d+]C\[\d+]$', address):
        row, col = address.split('C')
        row = row[2:-1]
        col = col[2:-1]
    else:
        raise Exception('Invalid address format ' + address)

    return sheet, col, row

def create_range(sheet, start_col, start_row, end_col, end_row):
    if not start_col or not end_col:
        raise Exception("Columns can't be None, start_col: [%s], start_row: [%s], end_col: [%s], "
                        "end_row: [%s]" % (start_col, start_row, end_col, end_row))

    result = start_col
    if start_row is not None:
        result += str(start_row)
    result += ':' + end_col
    if end_row is not None:
        result += str(end_row)
    if sheet is not None:
        if not sheet.find('!') > 0:
            sheet += '!'
        result = sheet + result
    return result

def resolve_function(formula):
    if formula and formula.startswith('='):
        fn, r = formula.split("(")
        return fn[1:]

def find_key_by_nested_key_value(dicts, deps, ky, value):
    for dep in deps:
        node = dicts[dep]
        if ky in node and value == node[ky]:
            return dep
    return None

def resolve_range_vlookup(rng, sheet, rng_node, formula_node, parent_ast, flatten=False):
    if not is_range(rng):
        return resolve_range_simple(rng, sheet)

    sheet = safe_str(create_sheet_plus(rng, sheet))
    sh, start, end = split_range(rng)
    sh, start_col, start_row = split_address(start)
    sh, end_col, end_row = split_address(end)
    search = find_key_by_nested_key_value(parent_ast.node, parent_ast.pred[formula_node], 'pos', 2)
    try:
        col = num2col(col2num(start_col) + int(search.tvalue) - 1)
    except Exception as ex:
        if not search:
            raise Exception("Search value not-found/incorrect, search: [%s], range: [%s], formula: "
                            "[%s]" % (search, rng, formula_node))
        # do it the normal way
        if logger.isEnabledFor(logging.INFO):
            logger.info("Unable to restrict the vlookup range: [%s], formula node: [%s]", rng, formula_node)

        return resolve_range_simple(rng, sheet)

    # Adding the search col
    first = create_range(sheet, start_col, start_row, start_col, end_row)
    addr1, nrows, ncols, rg = resolve_range(first, sheet=sheet, flatten=flatten)

    # Adding the result col
    second = create_range(sheet, col, start_row, col, end_row)
    addr2, nrows, ncols, rg = resolve_range(second, sheet=sheet, flatten=flatten)

    if logger.isEnabledFor(logging.DEBUG):
        logger.debug("Vlookup range: [%s] optimized to: [%s]", rng, [first, second])

    addr1 = list(zip(addr1, addr2))
    if flatten:
        addr1 = flatten(addr1)
        return addr1, 1, len(addr1)

    return addr1, nrows, 2, [first, second]

def resolve_range_hlookup(rng, sheet, rng_node, formula_node, parent_ast, flatten):
    if not is_range(rng):
        return resolve_range_simple(rng, sheet)

    sheet = safe_str(create_sheet_plus(rng, sheet))
    sh, start, end = split_range(rng)
    sh, start_col, start_row = split_address(start)
    sh, end_col, end_row = split_address(end)
    search = find_key_by_nested_key_value(parent_ast.node, parent_ast.pred[formula_node], 'pos', 2)
    try:
        row = int(start_row) + int(search.tvalue) - 1
    except Exception as ex:
        if not search:
            raise Exception("Search value not-found/incorrect, search: [%s], range: [%s], formula: "
                            "[%s]" % (search, rng, formula_node))
        # do it the normal way
        if logger.isEnabledFor(logging.INFO):
            logger.info("Unable to restrict the hlookup range: [%s], formula node: [%s]", rng, formula_node)
        return resolve_range_simple(rng, sheet)

    # Adding the search row
    first = create_range(sheet, start_col, start_row, end_col, start_row)
    addr1, nrows, ncols, rg = resolve_range(first, sheet=sheet, flatten=flatten)

    # Adding the second/result row
    second = create_range(sheet, start_col, row, end_col, row)
    addr2, nrows, ncols, rg = resolve_range(second, sheet=sheet, flatten=flatten)

    if logger.isEnabledFor(logging.DEBUG):
        logger.debug("Hlookup range: [%s] optimized to: [%s]", rng, [first, second])

    if flatten:
        addr1 = flatten([addr1, addr2])
        return addr1, 1, len(addr1)
    return [addr1, addr2], 2, ncols, [first, second]

def resolve_range(rng, sheet=None, flatten=False, parent_ast=None):
    # single cell, no range
    if not is_range(rng):
        return [sheet + rng], 1, 1

    if parent_ast:
        # if rng.find('!') > 0:
        #     s, rg = rng.split("!")
        # We assume that the specified range is the same as used as key in the ast graph. In case the sheet is missing
        # and is present in the ast graph, we can't find it.
        rng_node = find_range_node(parent_ast, rng)
        if rng_node:
            formula_node = find_formula_node(rng_node, parent_ast)
            if formula_node:
                if formula_node.tvalue.lower() == 'vlookup':
                    return resolve_range_vlookup(rng, sheet, rng_node, formula_node, parent_ast, flatten)
                elif formula_node.tvalue.lower() == 'hlookup':
                    return resolve_range_hlookup(rng, sheet, rng_node, formula_node, parent_ast, flatten)
    return resolve_range_simple(rng, sheet, flatten)

def find_range_node(parent_ast, rg):
    from .excel_node import RangeNode
    for node in parent_ast.nodes():
        if isinstance(node, RangeNode):
            if node.tvalue.replace('$', '') == rg:
                return node

                # nodes = [x for x in parent_ast.nodes() if isinstance(x, RangeNode) and x.tvalue == rg]
                # return nodes[0] if len(nodes) else None

def find_formula_node(rng_node, parent_ast):
    formula_node = parent_ast.succ[rng_node]
    if len(formula_node) > 1:
        raise Exception(
            "Only one formula expected: [%s], range: [%s], parent_ast: [%s]" % (formula_node, rng_node, parent_ast))

    return next(iter(formula_node))  # retrieve the first value in the dict

def resolve_range_simple(rng, sheet, flatten=False):
    sh, start, end = split_range(rng)
    sh, start_col, start_row = split_address(start)
    sh, end_col, end_row = split_address(end)

    if not start_row or not end_row or not start_col or not end_col:
        # As this point we assume the ranges are corrected such that the range contains a full start/end row with the
        # row number. Example: range "E:G" should be corrected to "E1:G234" using the max row/columns of a sheet.
        raise Exception(
            "Row/Col start/end missing, range: [%s], start row: [%s], end row: [%s], start col: [%s], end col: [%s]"
            % (rng, start_row, end_row, start_col, end_col))

    # Let's return some "marker" values we use to recognize the shorthand notition when required.
    # if start_row is None and end_row is None:
    #     return 'all', -1, -1

    start_row = int(start_row)
    end_row = int(end_row)

    if start_col == end_col:  # single column
        return resolve_range_col(rng, start_row, end_row, start_col, sheet)
    elif start_row == end_row:  # single row
        return resolve_range_row(rng, start_row, end_row, start_col, end_col, sheet)
    else:  # rectangular range
        return resolve_range_rectangle(rng, start_row, end_row, start_col, end_col, sheet, flatten)

def resolve_range_rectangle(rng, start_row, end_row, start_col, end_col, sheet, flatten):
    sheet = create_sheet_plus(rng, sheet, True)
    start_col_nr = col2num(start_col)
    end_col_nr = col2num(end_col)
    cells = []
    for r in range(start_row, end_row + 1):
        row = []
        for c in range(start_col_nr, end_col_nr + 1):
            row.append(safe_str(sheet) + num2col(c) + str(r))
        cells.append(row)

    if flatten:
        # flatten into one list
        l = flatten(cells)
        return l, 1, len(l), rng
    else:
        return cells, len(cells), len(cells[0]), rng

def resolve_range_row(rng, start_row, end_row, start_col, end_col, sheet):
    sheet = create_sheet_plus(rng, sheet, True)
    start_col_nr = col2num(start_col)
    end_col_nr = col2num(end_col)
    ncols = end_col_nr - start_col_nr + 1
    data = ["%s%s%s" % (s, num2col(c), r) for (s, c, r) in
            zip([safe_str(sheet)] * ncols, list(range(start_col_nr, end_col_nr + 1)), [start_row] * ncols)]
    return data, 1, len(data), rng

def resolve_range_col(rng, start_row, end_row, start_col, sheet):
    sheet = create_sheet_plus(rng, sheet, True)
    nrows = end_row - start_row + 1
    data = ["%s%s%s" % (s, c, r) for (s, c, r) in
            zip([safe_str(sheet)] * nrows, [start_col] * nrows, list(range(start_row, end_row + 1)))]
    return data, len(data), 1, rng

def col2num(col):
    if not col:
        raise Exception("Column may not be empty")

    tot = 0
    for i, c in enumerate([c for c in col[::-1] if c != "$"]):
        if c == '$':
            continue
        tot += (ord(c) - 64) * 26 ** i
    return tot

# convert back
def num2col(num):
    if num < 1:
        raise Exception("Number must be larger than 0: %s" % num)

    s = ''
    q = num
    while q > 0:
        (q, r) = divmod(q, 26)
        if r == 0:
            q = q - 1
            r = 26
        s = string.ascii_uppercase[int(r - 1)] + s
    return s

def address2index(a):
    sh, c, r = split_address(a)
    return col2num(c), int(r)

def index2addres(c, r, sheet=None):
    return "%s%s%s" % (sheet + "!" if sheet else "", num2col(c), r)

def get_linest_degree(excel, cl):
    # TODO: assumes a row or column of linest formulas & that all coefficients are needed

    sh, c, r, ci = cl.address_parts()
    # figure out where we are in the row

    # to the left
    i = ci - 1
    while i > 0:
        f = excel.get_formula_from_range(index2addres(i, r))
        if f is None or f != cl.formula:
            break
        else:
            i = i - 1

    # to the right
    j = ci + 1
    while True:
        f = excel.get_formula_from_range(index2addres(j, r))
        if f is None or f != cl.formula:
            break
        else:
            j = j + 1

    # assume the degree is the number of linest's
    degree = (j - i - 1) - 1  # last -1 is because an n degree polynomial has n+1 coefs

    # which coef are we (left most coef is the coef for the highest power)
    coef = ci - i

    # no linests left or right, try looking up/down
    if degree == 0:
        # up
        i = r - 1
        while i > 0:
            f = excel.get_formula_from_range("%s%s" % (c, i))
            if f is None or f != cl.formula:
                break
            else:
                i = i - 1

        # down
        j = r + 1
        while True:
            f = excel.get_formula_from_range("%s%s" % (c, j))
            if f is None or f != cl.formula:
                break
            else:
                j = j + 1

        degree = (j - i - 1) - 1
        coef = r - i

    # if degree is zero -> only one linest formula -> linear regression -> degree should be one
    return max(degree, 1), coef

def flatten(l):
    if is_collection(l):
        for el in l:
            if isinstance(el, collections.Iterable) and not isinstance(el, str):
                for sub in flatten(el):
                    yield sub
            else:
                yield el
    return l

def uniqueify(seq):
    seen = set()
    seen_add = seen.add
    return [x for x in seq if x not in seen and not seen_add(x)]

def is_valid(value, criteria):
    check = criteria_parser(criteria)
    return check(value)

def is_collection(args):
    return isinstance(args, (list, tuple))

def is_nr(value):
    return value is not None and isinstance(value, (int, float))

def is_number(s):  # http://stackoverflow.com/questions/354038/how-do-i-check-if-a-string-is-a-number-float-in-python
    try:
        float(s)
        return True
    except ValueError:
        return False

def is_leap_year(year):
    if not is_number(year):
        raise TypeError("%s must be a number" % str(year))
    if year <= 0:
        raise TypeError("%s must be strictly positive" % str(year))

    # Watch out, 1900 is a leap according to Excel => https://support.microsoft.com/en-us/kb/214326
    return (year % 4 == 0 and year % 100 != 0 or year % 400 == 0) or year == 1900

def get_max_days_in_month(month, year):
    if not is_number(year) or not is_number(month):
        raise TypeError("All inputs must be a number")
    if year <= 0 or month <= 0:
        raise TypeError("All inputs must be strictly positive")

    if month in (4, 6, 9, 11):
        return 30
    elif month == 2:
        if is_leap_year(year):
            return 29
        else:
            return 28
    else:
        return 31

def normalize_year(y, m, d):
    if m <= 0:
        y -= int(abs(m) / 12 + 1)
        m = 12 - (abs(m) % 12)
        normalize_year(y, m, d)
    elif m > 12:
        y += int(m / 12)
        m = m % 12

    if d <= 0:
        d += get_max_days_in_month(m, y)
        m -= 1
        y, m, d = normalize_year(y, m, d)

    else:
        if m in (4, 6, 9, 11) and d > 30:
            m += 1
            d -= 30
            y, m, d = normalize_year(y, m, d)
        elif m == 2:
            if (is_leap_year(y)) and d > 29:
                m += 1
                d -= 29
                y, m, d = normalize_year(y, m, d)
            elif (not is_leap_year(y)) and d > 28:
                m += 1
                d -= 28
                y, m, d = normalize_year(y, m, d)
        elif d > 31:
            m += 1
            d -= 31
            y, m, d = normalize_year(y, m, d)

    return y, m, d

def date_from_int(nb):
    if not is_number(nb):
        raise TypeError("%s is not a number" % str(nb))

    # origin of the Excel date system
    current_year = 1900
    current_month = 0
    current_day = 0

    while (nb > 0):
        if not is_leap_year(current_year) and nb > 365:
            current_year += 1
            nb -= 365
        elif is_leap_year(current_year) and nb > 366:
            current_year += 1
            nb -= 366
        else:
            current_month += 1
            max_days = get_max_days_in_month(current_month, current_year)

            if nb > max_days:
                nb -= max_days
            else:
                current_day = nb
                nb = 0

    return current_year, current_month, current_day

def criteria_parser(criteria):
    if is_number(criteria):
        def check(x):
            return x == criteria  # and type(x) == type(criteria)
    elif type(criteria) == str:
        search = re.search('(\W*)(.*)', criteria.lower()).group
        operator = search(1)
        value = search(2)
        value = float(value) if is_number(value) else str(value)

        if operator == '<':
            def check(x):
                if not is_number(x):
                    return False  # Added such that it skips non-numbers in for example the sumif function
                    # raise TypeError(
                    #     'excellib.countif() doesnt\'t work for checking non number items against non equality')
                return x < value
        elif operator == '>':
            def check(x):
                if not is_number(x):
                    return False
                    # raise TypeError(
                    #     'excellib.countif() doesnt\'t work for checking non number items against non equality')
                return x > value
        elif operator == '>=':
            def check(x):
                if not is_number(x):
                    return False
                    # raise TypeError(
                    #     'excellib.countif() doesnt\'t work for checking non number items against non equality')
                return x >= value
        elif operator == '<=':
            def check(x):
                if not is_number(x):
                    return False
                    # raise TypeError(
                    #     'excellib.countif() doesnt\'t work for checking non number items against non equality')
                return x <= value
        elif operator == '<>':
            def check(x):
                if not is_number(x):
                    return False
                    # raise TypeError(
                    #     'excellib.countif() doesnt\'t work for checking non number items against non equality')
                return x != value
        else:
            def check(x):
                return x == criteria
    else:
        raise Exception('Could\'t parse criteria %s' % criteria)

    return check

def find_corresponding_index(rng, criteria):
    # parse criteria
    check = criteria_parser(criteria)
    valid = []
    for index, item in enumerate(rng):
        if check(item) is not None:
            valid.append(index)

    return valid

def correct_numbers(val1, val2):
    try:
        if val1 == '':
            val1 = None
        elif val2 == '':
            val2 = None
        elif isinstance(val1, str):
            val1 = float(val1)
        elif isinstance(val2, str):
            val2 = float(val2)

        return val1, val2

    except Exception as ex:
        if val1 or val2:
            return None, None
        raise Exception("Incorrect values for val1=[%s], val2=[%s]" % (val1, val2)) from ex


if __name__ == '__main__':
    pass
