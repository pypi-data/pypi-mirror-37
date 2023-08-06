##region Setttings
sInputFolderpath = "../res/Input"
bPause = True
##endregion
##region Imports
import os
import FormatRosterData as FRD
import openpyxl
import TM_CommonPy as TM
import traceback
##endregion

def Main():
    with TM.WorkspaceContext("Output",bCDInto=True,bPreDelete=True):
        iTotalErrorFileCount = 0
        cWorkbooksToReformat = [] #Expects value to be tuple(vOldWorkbook,sFileName)
        #--- Write cNameToURL.txt
        print("  Getting NameToURL list..")
        cNameToURL = FRD.GetDict_NameToURL('http://www.espn.com/mens-college-basketball/team/roster/_/id/120')
        TM.Delete("cNameToURL.txt")
        with open('cNameToURL.txt','w') as vFile:
            for vKey, vValue in cNameToURL.items():
                vFile.write(vKey + " : " + vValue + "\n")
        #---Get OldWorkbooks
        print("  Gathering unformatted sheets..")
        for sFileName in os.listdir(sInputFolderpath):
            if (not sFileName.split(".")[-1] in ["xlsx","txt"]) or "~$" in sFileName or "template" in sFileName.lower():
                print("sFileName(IGNORED): "+sFileName)
                continue
            elif sFileName.split(".")[-1] == "xlsx":
                sFilePath = "../res/Input/"+sFileName
                vOldWorkbook = openpyxl.load_workbook(sFilePath)
                cWorkbooksToReformat.append((vOldWorkbook,sFileName))
            elif sFileName.split(".")[-1] == "txt":
                with open(os.path.join(sInputFolderpath,sFileName),'r') as vTextFile:
                    for sLine in vTextFile.readlines():
                        sLine = sLine.rstrip('\n') #probs a better way to do this.
                        if not sLine:
                            continue
                        for vKey in cNameToURL.keys():
                            if sLine.lower() in vKey.lower():
                                print(sFileName+" -  MATCHED:"+sLine+"("+vKey+")")
                                vOldWorkbook = FRD.GetWorkbook("http://www.espn.com"+cNameToURL[vKey])
                                sScrapedFileName = "Scraped_"+FRD.GetTitle("http://www.espn.com"+cNameToURL[vKey])
                                cWorkbooksToReformat.append((vOldWorkbook,sScrapedFileName))
                                break
                        else:
                            print(sFileName+" - Could not match:"+sLine)
            else:
                iTotalErrorFileCount += 1
                print("**ERROR:Could not get workbook from sFileName:"+sFileName)
                continue
        #---Create NewWorkbooks
        print("  Creating formatted sheets..")
        for vOldWorkbook, sFileName in cWorkbooksToReformat:
            print("OldFileName:"+sFileName)
            #---Edit
            vOldSheet = vOldWorkbook.active
            vNewWorkbook = openpyxl.Workbook()
            vNewSheet = vNewWorkbook.active
            bSuccess = True
            bSuccess &= FRD.FormatName(vOldSheet,vNewSheet)
            bSuccess &= FRD.FormatHometown(vOldSheet,vNewSheet)
            bSuccess &= FRD.FormatHeight(vOldSheet,vNewSheet)
            if not "women" in sFileName.lower():
                bSuccess &= FRD.FormatWeight(vOldSheet,vNewSheet)
            bSuccess &= FRD.FormatSchoolyear(vOldSheet,vNewSheet)
            bSuccess &= FRD.FormatPosition(vOldSheet,vNewSheet)
            bSuccess &= FRD.AppendOldSheet(vOldSheet,vNewSheet)
            #---Save
            if not bSuccess:
                print("New_FileName:"+sFileName.split(".")[0]+"_Reformatted(ERRORS).xlsx")
                vNewWorkbook.save(sFileName.split(".")[0]+"_Reformatted(ERRORS).xlsx")
                iTotalErrorFileCount += 1
            else:
                print("New_FileName:"+sFileName.split(".")[0]+"_Reformatted.xlsx")
                vNewWorkbook.save(sFileName.split(".")[0]+"_Reformatted.xlsx")
    if iTotalErrorFileCount:
        print("TOTAL ERROR FILES`iTotalErrorFileCount:"+str(iTotalErrorFileCount))
    else:
        print("Success`iTotalErrorFileCount:"+str(iTotalErrorFileCount))

try:
    Main()
except PermissionError:
    print("PERMISSION_ERROR\n\tI'd recommend to just try again.\n\tOtherwise, close all extra programs and then retry.")
    TM.DisplayDone()
except Exception as e:
    TM.DisplayException(e)
else:
    if bPause:
        TM.DisplayDone()
