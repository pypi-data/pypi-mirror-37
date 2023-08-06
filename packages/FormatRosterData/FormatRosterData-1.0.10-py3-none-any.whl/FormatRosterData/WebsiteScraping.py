import os
import openpyxl
import TM_CommonPy as TM
import lxml.html
import requests

def DigForText(vElem):
    for i in range(15):
        if not vElem.text is None:
            break;
        vElem = vElem[0]
    else:
        return ""
    return vElem.text

def GetDict_NameToURL_Men():
    sURL = 'http://www.espn.com/mens-college-basketball/team/roster/_/id/120'
    #---Get vList
    vRosterPage = requests.get(sURL)
    tree = lxml.html.fromstring(vRosterPage.content)
    #vList = tree.xpath('//*[@id="fittPageContainer"]/div[3]/div[2]/div[1]/div[1]/section/section/div[1]/div/select[1]')[0]
    vList = tree.xpath('//select[@class="dropdown__select"]')[1] #Perhaps there is a more reliable xpath
    #---
    cNameToURL = dict()
    for vItem in vList:
        if "More NCAAM teams" == DigForText(vItem): #Irrelevant key and value
            continue
        cNameToURL[DigForText(vItem)] = "http://www.espn.com" + vItem.attrib['data-url']
    return cNameToURL

def GetDict_NameToURL_Women():
    sURL = 'http://www.espn.com/womens-college-basketball/team/roster/_/id/120'
    #---Get vList
    vRosterPage = requests.get(sURL)
    tree = lxml.html.fromstring(vRosterPage.content)
    #vList = tree.xpath('//*[@id="fittPageContainer"]/div[3]/div[2]/div[1]/div[1]/section/section/div[1]/div/select[1]')[0]
    vList = tree.xpath('//select[@class="select-box"]')[0] #Perhaps there is a more reliable xpath
    #---
    cNameToURL = dict()
    for vItem in vList:
        if "More NCAAM teams" == DigForText(vItem): #Irrelevant key and value
            continue
        cNameToURL[DigForText(vItem)] = "http:" + vItem.attrib['value']
    return cNameToURL

def GetTitle(sURL):
    vRosterPage = requests.get(sURL)
    tree = lxml.html.fromstring(vRosterPage.content)
    if "women" in sURL.lower():
        #---Get vRosterTitle
        vRosterTitle = tree.xpath('//h1[@class="h2"]')[0]
    else:
        #---Get vRosterTitle
        vRosterTitle = tree.xpath('//h1[@class="headline__h1 dib"]')[0]
    if vRosterTitle is None:
        print("vRosterTitle is None")
        raise
    return DigForText(vRosterTitle).replace(" ","")

def GetWorkbook_Men(sURL):
    #---Get RosterTable
    vRosterPage = requests.get(sURL)
    tree = lxml.html.fromstring(vRosterPage.content)
    vRosterTableHeader = tree.xpath('//thead[@class="Table2__sub-header Table2__thead"]')[0]
    vRosterTable = tree.xpath('//tbody[@class="Table2__tbody"]')[0]
    #---Convert vRosterTable to openpyxl doc
    vWorkbook = openpyxl.Workbook()
    vSheet = vWorkbook.active
    for iCol, vItem in enumerate(vRosterTableHeader[0]):
        vSheet[openpyxl.utils.get_column_letter(iCol+1)+str(1)] = DigForText(vItem)
    for iRow, vRow in enumerate(vRosterTable):
        for iCol, vItem in enumerate(vRow):
            vSheet[openpyxl.utils.get_column_letter(iCol+1)+str(iRow+1+1)] = DigForText(vItem) #xlsx iCol and iRow start index at 1. Row gets another +1 for header.
    return vWorkbook

def GetWorkbook_Women(sURL):
    #---Get RosterTable
    vRosterPage = requests.get(sURL)
    tree = lxml.html.fromstring(vRosterPage.content)
    vRosterTable = tree.xpath('//table[@class="tablehead"]')[0]
    #---Convert vRosterTable to openpyxl doc
    vWorkbook = openpyxl.Workbook()
    vSheet = vWorkbook.active
    for iCol, vItem in enumerate(vRosterTable[1]): #vRosterTable[1] is header row
        vSheet[openpyxl.utils.get_column_letter(iCol+1)+str(1)] = DigForText(vItem)
    for iRow, vRow in enumerate(vRosterTable):
        if iRow < 2: #vRosterTable[2] begins data rows
            continue
        iRow = iRow - 2 #to keep consistency, why not?
        for iCol, vItem in enumerate(vRow):
            vSheet[openpyxl.utils.get_column_letter(iCol+1)+str(iRow+1+1)] = DigForText(vItem) #xlsx iCol and iRow start index at 1. Row gets another +1 for header.
    return vWorkbook
