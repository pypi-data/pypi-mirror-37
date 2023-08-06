import os
import openpyxl
import TM_CommonPy as TM

def GetPos(vCell,iColAdjustment=0):
    return openpyxl.utils.get_column_letter(openpyxl.utils.column_index_from_string(vCell.column)+iColAdjustment)+str(vCell.row)

def SplitName(vOldSheet,vNewSheet):
    bSuccess = True
    #---Determine Name Column and Row
    for vCell in (vOldSheet['1']+vOldSheet['2']):
        try:
            if "name" in vCell.value.lower():
                iRow = vCell.row
                sColumn = vCell.column
                break
        except (TypeError, AttributeError):
            pass
    else:
        print("**ERROR:Could not find Name header")
        return False
    #---
    iMaxCol = len(vNewSheet['1'])
    for vCell in vOldSheet['B']:
        #-Skip past header
        if vCell.row <= iRow:
            continue
        #-
        cSplitString = vCell.value.strip().split(None,1) #split None splits at first whitespace, a necessary bugfix
        try:
            vNewSheet[openpyxl.utils.get_column_letter(iMaxCol+1)+str(vCell.row)] = cSplitString[0]
            vNewSheet[openpyxl.utils.get_column_letter(iMaxCol+2)+str(vCell.row)] = cSplitString[1]
        except IndexError:
            bSuccess = False
    return bSuccess

def SplitTown(vOldSheet,vNewSheet):
    bSuccess = True
    #---Determine Name Column and Row
    for vCell in (vOldSheet['1']+vOldSheet['2']):
        try:
            if "hometown" in vCell.value.lower():
                iRow = vCell.row
                sColumn = vCell.column
                break
        except (TypeError, AttributeError):
            pass
    else:
        print("**ERROR:Could not find Hometown header")
        return False
    #---
    iMaxCol = len(vNewSheet['1'])
    for vCell in vOldSheet[sColumn]:
        #-Skip past header
        if vCell.row <= iRow:
            continue
        #-
        cSplitString = vCell.value.split(", ")
        try:
            vNewSheet[openpyxl.utils.get_column_letter(iMaxCol+1)+str(vCell.row)] = cSplitString[0]
            vNewSheet[openpyxl.utils.get_column_letter(iMaxCol+2)+str(vCell.row)] = cSplitString[1].split("/")[0].strip()
        except:
            bSuccess=False
            raise
    return bSuccess

def ConvertDateToHeight(vDate):
    #---Filter
    if vDate is None:
        return ""
    #---
    cTemp = str(vDate).split("-")
    return int(cTemp[1])*12+int(cTemp[2].split(None)[0])

def TranslateHeight(vOldSheet,vNewSheet):
    bSuccess = True
    #---Determine Height Col and Row
    for vCell in (vOldSheet['1']+vOldSheet['2']):
        try:
            if "ht." in vCell.value.lower() or "height" in vCell.value.lower():
                iRow = vCell.row
                sColumn = vCell.column
                break
        except (TypeError, AttributeError):
            pass
    else:
        print("**ERROR:Could not find Height header")
        return False
    #---
    iMaxCol = len(vNewSheet['1'])
    for vCell in vOldSheet[sColumn]:
        #-Skip past header
        if vCell.row <= iRow:
            continue
        #-
        vNewSheet[openpyxl.utils.get_column_letter(iMaxCol+1)+str(vCell.row)] = ConvertDateToHeight(vCell.value)
    return bSuccess

def GetWeight(vOldSheet,vNewSheet):
    bSuccess = True
    #---Determine Weight Col and Row
    for vCell in (vOldSheet['1']+vOldSheet['2']):
        try:
            if "wt." in vCell.value.lower() or "weight" in vCell.value.lower():
                iRow = vCell.row
                sColumn = vCell.column
                break
        except (TypeError, AttributeError):
            pass
    else:
        print("""**ERROR:Could not find Weight header.
            \n\tIf its a Women's excel sheet, you can rename the excel sheet to
            \n\tinclude the word \'Women\' so that this program doesn't try to
            \n\tfind the weight column.""")
        return False
    #---
    iMaxCol = len(vNewSheet['1'])
    for vCell in vOldSheet[sColumn]:
        #-Skip past header
        if vCell.row <= iRow:
            continue
        #-
        vNewSheet[openpyxl.utils.get_column_letter(iMaxCol+1)+str(vCell.row)] = vCell.value
    return bSuccess

def ConvertFshSophJrSenToInt(vValue):
    if "Fr." in vValue or "freshman" in vValue.lower():
        return 1
    elif "So." in vValue or "sophmore" in vValue.lower() or "soph" in vValue.lower():
        return 2
    elif "Jr." in vValue or "junior" in vValue.lower():
        return 3
    elif "Sr." in vValue or "senior" in vValue.lower() or "gr." in vValue.lower():
        return 4
    else:
        return 0

def GetSchoolyear(vOldSheet,vNewSheet):
    bSuccess = True
    #---Determine GetSchoolyear Col and Row
    for vCell in (vOldSheet['1']+vOldSheet['2']):
        try:
            if "cl." in vCell.value.lower() or "year" in vCell.value.lower() or "yr." in vCell.value.lower():
                iRow = vCell.row
                sColumn = vCell.column
                break
        except (TypeError, AttributeError):
            pass
    else:
        print("**ERROR:Could not find Schoolyear header")
        return False
    #---
    iMaxCol = len(vNewSheet['1'])
    for vCell in vOldSheet[sColumn]:
        #-Skip past header
        if vCell.row <= iRow:
            continue
        #-
        iInt = ConvertFshSophJrSenToInt(vCell.value)
        if iInt == 0:
            print("**ERROR:Could not translate FshSophJrSen number from:"+str(vCell.value))
            bSuccess = False
        vNewSheet[openpyxl.utils.get_column_letter(iMaxCol+1)+str(vCell.row)] = iInt
    return bSuccess

def AppendOldSheet(vOldSheet,vNewSheet):
    iMaxCol = len(vNewSheet['1'])
    bSuccess = True
    try:
        for cColumn in vOldSheet.iter_cols():
            for vCell in cColumn:
                vNewSheet[GetPos(vCell,iColAdjustment=iMaxCol+1)] = vCell.value
    except:
        print("**ERROR:Could not append old sheet")
        bSuccess=False
    return bSuccess
