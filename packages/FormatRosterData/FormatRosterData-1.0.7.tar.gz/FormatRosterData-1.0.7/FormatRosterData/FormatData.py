import os
import openpyxl
import TM_CommonPy as TM

def IsEmptySheet(vSheet):
    for cRows in vSheet.iter_rows():
        for vCell in cRows:
            if not vCell.value is None:
                return False
    return True

def GetPos(vCell,iColAdjustment=0):
    return openpyxl.utils.get_column_letter(openpyxl.utils.column_index_from_string(vCell.column)+iColAdjustment)+str(vCell.row)

def FormatName(vOldSheet,vNewSheet):
    bSuccess = True
    #---Determine Name Header Column and Row
    for vCell in (vOldSheet['1']+vOldSheet['2']):
        try:
            if "name" in vCell.value.lower():
                iHeaderRow = vCell.row
                sColumn = vCell.column
                break
        except (TypeError, AttributeError):
            pass
    else:
        print("**ERROR:Could not find \'Name\' header")
        return False
    #---Determine iPrevMaxCol
    if IsEmptySheet(vNewSheet):
        iPrevMaxCol = 0
    else:
        iPrevMaxCol = len(vNewSheet['1'])
    #---
    for vCell in vOldSheet['B']:
        #-header
        if vCell.row < iHeaderRow:
            continue
        elif vCell.row == iHeaderRow:
            vNewSheet[openpyxl.utils.get_column_letter(iPrevMaxCol+1)+str(vCell.row)] = "Name"
            continue
        #-
        cSplitString = vCell.value.strip().split(None,1) #split None splits at first whitespace, a necessary bugfix
        try:
            vNewSheet[openpyxl.utils.get_column_letter(iPrevMaxCol+1)+str(vCell.row)] = cSplitString[0]
            vNewSheet[openpyxl.utils.get_column_letter(iPrevMaxCol+2)+str(vCell.row)] = cSplitString[1]
        except IndexError:
            bSuccess = False
    return bSuccess

def FormatHometown(vOldSheet,vNewSheet):
    bSuccess = True
    #---Determine Town Header Column and Row
    for vCell in (vOldSheet['1']+vOldSheet['2']):
        try:
            if "hometown" in vCell.value.lower():
                iHeaderRow = vCell.row
                sColumn = vCell.column
                break
        except (TypeError, AttributeError):
            pass
    else:
        print("**ERROR:Could not find \'Hometown\' header")
        return False
    #---Determine iPrevMaxCol
    if IsEmptySheet(vNewSheet):
        iPrevMaxCol = 0
    else:
        iPrevMaxCol = len(vNewSheet['1'])
    #---
    for vCell in vOldSheet[sColumn]:
        #-header
        if vCell.row < iHeaderRow:
            continue
        elif vCell.row == iHeaderRow:
            vNewSheet[openpyxl.utils.get_column_letter(iPrevMaxCol+1)+str(vCell.row)] = "Hometown"
            continue
        #-
        cSplitString = vCell.value.split(", ")
        try:
            vNewSheet[openpyxl.utils.get_column_letter(iPrevMaxCol+1)+str(vCell.row)] = cSplitString[0]
            vNewSheet[openpyxl.utils.get_column_letter(iPrevMaxCol+2)+str(vCell.row)] = cSplitString[1].split("/")[0].strip()
        except:
            bSuccess=False
            raise
    return bSuccess

def ConvertDateToHeight(vDate):
    #---Filter
    if vDate is None:
        return ""
    #---
    cTemp = str(vDate).split("-")
    return int(cTemp[1])*12+int(cTemp[2].split(None)[0])

def ConvertHeightStrToHeight(vHeightStr):
    #---Filter
    if vHeightStr is None:
        return ""
    #---
    cNums = TM.GetNumsInString(vHeightStr)
    if len(cNums) != 2:
        print("**ERROR:Could not get height number from:"+str(vHeightStr))
        return
    return cNums[0]*12+cNums[1]

def FormatHeight(vOldSheet,vNewSheet):
    bSuccess = True
    #---Determine Height Header Col and Row
    for vCell in (vOldSheet['1']+vOldSheet['2']):
        try:
            if "ht." in vCell.value.lower() or "height" in vCell.value.lower() or "ht" == vCell.value.lower():
                iHeaderRow = vCell.row
                sColumn = vCell.column
                break
        except (TypeError, AttributeError):
            pass
    else:
        print("**ERROR:Could not find \'Height\' header")
        return False
    #---Determine iPrevMaxCol
    if IsEmptySheet(vNewSheet):
        iPrevMaxCol = 0
    else:
        iPrevMaxCol = len(vNewSheet['1'])
    #---
    for vCell in vOldSheet[sColumn]:
        #-header
        if vCell.row < iHeaderRow:
            continue
        elif vCell.row == iHeaderRow:
            vNewSheet[openpyxl.utils.get_column_letter(iPrevMaxCol+1)+str(vCell.row)] = "Height"
            continue
        #-
        if vCell.value is None:
            pass
        elif "-" in str(vCell.value):
            vNewSheet[openpyxl.utils.get_column_letter(iPrevMaxCol+1)+str(vCell.row)] = ConvertDateToHeight(vCell.value)
        elif "\"" in str(vCell.value):
            iHeight = ConvertHeightStrToHeight(vCell.value)
            if iHeight is None:
                bSuccess = False
            else:
                vNewSheet[openpyxl.utils.get_column_letter(iPrevMaxCol+1)+str(vCell.row)] = iHeight
        else:
            print("**ERROR:Could not determine Height's format from:"+str(vCell.value))
            bSuccess = False
    return bSuccess

def FormatWeight(vOldSheet,vNewSheet):
    bSuccess = True
    #---Determine Weight Header Col and Row
    for vCell in (vOldSheet['1']+vOldSheet['2']):
        try:
            if "wt." in vCell.value.lower() or "weight" in vCell.value.lower() or "wt" == vCell.value.lower():
                iHeaderRow = vCell.row
                sColumn = vCell.column
                break
        except (TypeError, AttributeError):
            pass
    else:
        print("""**ERROR:Could not find \'Weight\' header.
            \tIf its a Women's excel sheet, you can rename the excel sheet to
            \tinclude the word \'Women\' so that this program doesn't try to
            \tfind the weight column.""")
        return False
    #---Determine iPrevMaxCol
    if IsEmptySheet(vNewSheet):
        iPrevMaxCol = 0
    else:
        iPrevMaxCol = len(vNewSheet['1'])
    #---
    for vCell in vOldSheet[sColumn]:
        #-header
        if vCell.row < iHeaderRow:
            continue
        elif vCell.row == iHeaderRow:
            vNewSheet[openpyxl.utils.get_column_letter(iPrevMaxCol+1)+str(vCell.row)] = "Weight"
            continue
        #-
        cNums = TM.GetNumsInString(str(vCell.value))
        if len(cNums) != 1:
            bSuccess=False
            print("**ERROR:Could not determine weight number from:"+str(vCell.value))
        vNewSheet[openpyxl.utils.get_column_letter(iPrevMaxCol+1)+str(vCell.row)] = cNums[0]
    return bSuccess

def ConvertFshSophJrSenToInt(vValue):
    if "fr." in vValue.lower() or "freshman" in vValue.lower() or "fr" == vValue.lower():
        return 1
    elif "so." in vValue.lower() or "sophmore" in vValue.lower() or "soph" in vValue.lower() or "so" == vValue.lower():
        return 2
    elif "jr." in vValue.lower() or "junior" in vValue.lower()  or "jr" == vValue.lower():
        return 3
    elif "sr." in vValue.lower() or "senior" in vValue.lower() or "gr." in vValue.lower() or "sn." in vValue.lower()  or "sr" == vValue.lower()  or "gr" == vValue.lower():
        return 4
    else:
        return

def FormatSchoolyear(vOldSheet,vNewSheet):
    bSuccess = True
    #---Determine GetSchoolyear Header Col and Row
    for vCell in (vOldSheet['1']+vOldSheet['2']):
        try:
            if "cl." in vCell.value.lower() or "class" in vCell.value.lower() or "year" in vCell.value.lower() or "yr." in vCell.value.lower():
                iHeaderRow = vCell.row
                sColumn = vCell.column
                break
        except (TypeError, AttributeError):
            pass
    else:
        print("**ERROR:Could not find \'Schoolyear\' header")
        return False
    #---Determine iPrevMaxCol
    if IsEmptySheet(vNewSheet):
        iPrevMaxCol = 0
    else:
        iPrevMaxCol = len(vNewSheet['1'])
    #---
    for vCell in vOldSheet[sColumn]:
        #-header
        if vCell.row < iHeaderRow:
            continue
        elif vCell.row == iHeaderRow:
            vNewSheet[openpyxl.utils.get_column_letter(iPrevMaxCol+1)+str(vCell.row)] = "Year"
            continue
        #-
        iInt = ConvertFshSophJrSenToInt(vCell.value)
        if iInt is None:
            print("**ERROR:Could not translate FshSophJrSen number from:"+str(vCell.value))
            bSuccess = False
        vNewSheet[openpyxl.utils.get_column_letter(iPrevMaxCol+1)+str(vCell.row)] = iInt
    return bSuccess


def FormatPosition(vOldSheet,vNewSheet):
    bSuccess = True
    #---Determine FormatPosition Header Col and Row
    for vCell in (vOldSheet['1']+vOldSheet['2']):
        try:
            if "pos" in vCell.value.lower():
                iHeaderRow = vCell.row
                sColumn = vCell.column
                break
        except (TypeError, AttributeError):
            pass
    else:
        print("**ERROR:Could not find \'Position\' header")
        return False
    #---Determine iPrevMaxCol
    if IsEmptySheet(vNewSheet):
        iPrevMaxCol = 0
    else:
        iPrevMaxCol = len(vNewSheet['1'])
    #---
    for vCell in vOldSheet[sColumn]:
        #-header
        if vCell.row < iHeaderRow:
            continue
        elif vCell.row == iHeaderRow:
            vNewSheet[openpyxl.utils.get_column_letter(iPrevMaxCol+1)+str(vCell.row)] = "Position"
            continue
        #-
        if vCell.value is None:
            pass
        else:
            sFormattedValue = ""
            for sString in str(vCell.value).split("/"):
                sFormattedValue += sString[0] + "/"
            sFormattedValue = sFormattedValue[:-1]
            vNewSheet[openpyxl.utils.get_column_letter(iPrevMaxCol+1)+str(vCell.row)] = sFormattedValue
    return bSuccess

def AppendOldSheet(vOldSheet,vNewSheet):
    iPrevMaxCol = len(vNewSheet['1'])
    bSuccess = True
    try:
        for cColumn in vOldSheet.iter_cols():
            for vCell in cColumn:
                vNewSheet[GetPos(vCell,iColAdjustment=iPrevMaxCol+2)] = vCell.value
    except:
        print("**ERROR:Could not append old sheet")
        bSuccess=False
    return bSuccess
