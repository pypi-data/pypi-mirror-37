'''
QuickGrid library - very simple communication with spreadsheets. 
v1
'''
import csv
import unicodecsv as ucsv
import os
from itertools import groupby
from collections import Counter
import codecs
import sys
import six
csv.field_size_limit(sys.maxsize)

try:
    from openpyxl import load_workbook
    import xlrd
    import xlwt
    from xlwt import Workbook
except:
    load_workbook = None
    xlrd = None
    xlwt = None
    Workbook = object

def make_safe(v):
    if isinstance(v,float):
        return str(v).lower().strip()
    if v:
        return v.lower().strip()
    else:
        return ""

class FlexiBook(Workbook):
    """
    modification to xlwt library - for merging multiple sheets
    """
    def __init__(self,*args,**kwargs):
        super(FlexiBook,self).__init__(*args,**kwargs)

    def add_sheet_from_ql(self,ql):

        sheet_name = ql.name
        if not sheet_name:
            sheet_name = "quicklist"
        ws = self.add_sheet(sheet_name,True)

        row = 0

        for i, h in enumerate(ql.header):
            ws.write(row, i, h)

        for r in ql.data:
            row += 1
            for iter, item in enumerate(r):
                ws.write(row, iter, item)

        return self
    
    def merge_qls(self,qls):
        for q in qls:
            self.add_sheet_from_ql(q)
            
        return self

        

def import_csv(s_file,unicode=False,start=0,limit=None,codec=""):
    """
    imports csvs, returns head/body - switch to use different csv processor
    """
    body = []
    if unicode:
        mod = ucsv
    else:
        mod = csv
    count = -1
    
    internal_limit = None
    if limit:
        internal_limit = start + limit
    
    if six.PY2:
        readtype = "rb"
    else:
        readtype = "rt"
    
    if codec:
        o = codecs.open
        args = [s_file, readtype,codec]
    else:
        o = open
        args = (s_file, readtype)
    
    with o(*args) as f:
        reader = mod.reader((line.replace('\0','') for line in f))
        for row in reader:
            count += 1
            if count == 0:
                yield row
                continue    
            if count >= start:
                yield row
            if limit and count == internal_limit + 1:
                break

def import_xls(s_file, tab="", header_row=0):
    """
    does what you'd expect
    """
    wb = xlrd.open_workbook(filename=s_file)
    if tab == "":
        tab = 0
    if type(tab) == int:
        ws = wb.sheet_by_index(tab)
    else:
        ws = wb.sheet_by_name(tab)  # ws is now an IterableWorksheet

    def generate_rows():
        num_cols = ws.ncols
        for r in range(0, ws.nrows):
            row = [ws.cell(r, c) for c in range(0, num_cols)]
            yield row


    for count, row in enumerate(generate_rows()):
        r = [x.value for x in row]
        if count >= header_row:
            yield r

def import_xlsx(s_file, tab="", header_row=0):
    """
    imports xlsx files
    """
    wb = load_workbook(filename=s_file, read_only=True, data_only=True)
    if tab == "":
        tab = wb.sheetnames[0]
    ws = wb[tab]  # ws is now an IterableWorksheet

    for count, row in enumerate(ws.rows):
        r = [x.value for x in row]
        if count >= header_row:
            yield r

def export_csv(file, header, body, force_unicode=False):
    """
    exports csv (non-unicode)
    """
    if force_unicode:
        module = ucsv
    else:
        module = csv
    with open(file, 'wb') as f:
        w = module.writer(f)
        w.writerows([header])
        w.writerows(body)

class ItemRow(list):
    """
    object returned while iterating through a quick list
    """   
    def _key_to_index(self, value):
        if isinstance(value,int) or isinstance(value,slice):
            return value
        r_v = value.strip().lower()
        try:
            lookup = self._header_dict[r_v]
        except KeyError:
            lookup = value
        return lookup
    
    def __init__(self,quick_grid_instance,header_dict,*args,**kwargs):
        
        
        self._header_dict =header_dict
        self._qg = quick_grid_instance
        
        super(ItemRow,self).__init__(*args,**kwargs)
        diff = len(self._qg.header) - len(self)
        if diff > 0:
            self.extend(["" for x in range(0,diff)])

    def __getitem__(self, key):
        return super(ItemRow, self).__getitem__(self._key_to_index(key))

    def __setitem__(self, key, value):
        return super(ItemRow, self).__setitem__(self._key_to_index(key), value)

    def __getattr__(self, attr):
        
        if  attr != "_header_dict" and attr in self._header_dict:
            return self[attr]
        return object.__getattribute__(self, attr)

    def __setattr__(self, attr, value):
        
        if attr != "_header_dict" and attr in self._header_dict and attr[0] != "_":
            self[attr] = value
        else:
            super(ItemRow,self).__setattr__(attr,value)

class QuickGrid(object):

    """
    A very simple files interface - loads files into memory so basic reads can be done. 
    
    """
    
    @classmethod
    def split_csv_count(cls,filename,save_folder,limit=1000,force_unicode=False):
    

        if force_unicode:
            mod = ucsv
        else:
            mod = csv
        
        header = cls()
        f = open(filename, 'rb')
        reader = mod.reader(f)
        count = -1
        for row in reader:
            count += 1
            if count == 0:
                header.header = row
                current = cls(count)
                current.header = header.header
                continue
        
            current.data.append(row)
            if count % limit == 0:
                current.save([save_folder,"{0}.csv".format(current.name)],force_unicode=force_unicode)
                current.data = []
                current.name = count
                
        current.save([save_folder,"{0}.psv".format(current.name)],force_unicode=force_unicode)
    
    @classmethod
    def split_csv(cls,filename,save_folder,column_id=0,unicode=False):
    
        if unicode:
            mod = ucsv
        else:
            mod = csv

        last_value = None
        current = None
        header = cls()
        f = open(filename, 'rb')
        reader = mod.reader(f)
        count = -1
        for row in reader:
            count += 1
            if count == 0:
                header.header = row
                if isinstance(column_id,str):
                    column_id = header.header_di()[make_safe(column_id)]
                continue
            if row[column_id] == None:
                continue
            if row[column_id] != last_value:
                if current:
                    current.save([save_folder,"{0}.csv".format(current.name)])
                current = cls(row[column_id])
                current.header = header.header
                last_value = row[column_id]
            current.data.append(row)
        current.save([save_folder,"{0}.csv".format(current.name)])


    @classmethod
    def merge(cls,to_merge):
        """
        join a bunch of QuickGrids together
        """
        new = cls()
        source = to_merge[0]
        new.data = list(source.data)
        new.header = list(source.header)
        new.add_qg(to_merge[1:])
        return new
    
    def count(self,*columns):
        
        count = self.__class__()
        count.header = list(columns) + ["count"]
        values = []
        for r in self:
            
            v = "|".join(["{0}".format(r[x]) for x in columns])
            values.append(v)
            
        counter = Counter(values)
        
        for k,v in counter.iteritems():
            row = k.split("|")
            row += [v]
            count.add(row)
            
        return count
        
 
        
    def add_qg(self,to_merge):
        for m in to_merge:
            if m.header == self.header:
                self.data.extend(m.data)
            else:
                raise ValueError("Header mismatch")
            
    
    def __init__(self, name="",header=[],cached=True):

        self.name = name
        self.header = header
        self.data = []
        self.filename = None
        self.cached = cached
        self.transformers = []

    def add(self,row):
        
        if isinstance(row,dict):
            #convert a dictionary into a new row object
            nrow = ["" for x in self.header]
            for k,v in row.iteritems():
                loc = self.col_to_location(k)
                nrow[loc] = v
            row = nrow

        try:
            assert len(row) == len(self.header)
        except Exception:
            raise ValueError("New row is different size than header.")
        self.data.append(row)

    def combine(self, *args, **kwargs):
        
        def combo(row):
            return "|".join(["{0}".format(row[c]) for c in args])
        
        if "name" in kwargs:
            name = kwargs["name"]
        else:
            name = "_".join(args)
        
        self.generate_col(name,combo)
        
    def generate_col(self,name,generate_function):
        self.header.append(name)
        for r in self:
            r[-1] = generate_function(r)
        
    def transform(self,col_name,function):
        if self.cached:
            for r in self:
                r[col_name] = function(r[col_name])
        else:
            self.transformers.append((col_name,function))
        
        

    def rename_column(self,old_name,new_name):
        """
        find old header - replace with enw one
        """
        safe_col = make_safe(old_name)
        di = self.header_di()
        location = di[safe_col]
        self.header[location] = new_name

        
    def sort(self,*cols):
        """
        sort according to columns entered (start with - to reverse)
        """
        for c in cols[::-1]:
            reverse = c[0] == "-"
            if reverse:
                v = c[1:]
            else:
                v = c
            self.data.sort(key = lambda x:x[self.col_to_location(v)],reverse=reverse)

    def split_on_unique(self,col,col_no=None):
        if col_no == None:
            location = self.col_to_location(col)
        else:
            location = col_no
        func = lambda x:x[location]
        self.data.sort(key = func)
        for k,g in groupby(self.data,func):
            if k:
                name = self.name + " {0}".format(k)
            else:
                name = self.name + " " + "None"
            n = self.__class__(name)
            n.header = list(self.header)
            n.data = list(g)
            yield k,n

    def counter(self,col):
        loc = self.col_to_location(col)
        return Counter([x[loc] for x in self.data])

    def col_to_location(self,col):
        def make_safe(v):
            if v:
                return v.lower().strip()
            else:
                return ""
            
        safe_col = make_safe(col)
        
        di = self.header_di()
        return  di[safe_col]        

    def include(self,col,values):
        """
        returns an generator of only rows where value in col
        """
        location = self.col_to_location(col)
        indexes = [x for x,y in enumerate(self.data) if y[location]in values]
        
        return self.__iter__(indexes=indexes)

    def only(self,col,value):
        """
        returns an generator of only rows where col = value
        """
        location = self.col_to_location(col)
        indexes = [x for x,y in enumerate(self.data) if y[location] == value]
        
        return self.__iter__(indexes=indexes)
        
    def drop(self,**kwargs):
        
        def make_safe(v):
            if v:
                return v.lower().strip()
            else:
                return ""
            
        di = self.header_di()
        loc = lambda x: di[make_safe(x)]

        internal = [[loc(k),v] for k,v in kwargs.iteritems()]
        new_data = []  
        
        for r in self.data:
            match = True
            for l, v in internal:
                if r[l] == v:
                    match = False
                    break
            if match:
                new_data.append(r)
        self.data = new_data
    
    def expand(self,col,seperator=","):
        safe_col = make_safe(col)
        di = self.header_di()
        location = di[safe_col]
        
        new_data = []
        
        for r in self.data:
            cell = r[location]
            if cell:
                values = cell.split(seperator)
            else:
                values = [""]
            for v in values:
                row = list(r)
                row[location] = v
                new_data.append(row)
        
        self.data = new_data
        
    def exclude(self,col,value):
        """
        returns an generator of only rows where col != value
        """
        def make_safe(v):
            if v:
                return v.lower().strip()
            else:
                return ""
            
        safe_col = make_safe(col)
        
        di = self.header_di()
        location = di[safe_col]
        indexes = [x for x,y in enumerate(self.data) if y[location] != value]
        
        return self.__iter__(indexes=indexes)
        


    def unique(self,col):
        """
        returns an generator of rows with unique values in col
        """
        def make_safe(v):
            if v:
                return v.lower().strip()
            else:
                return ""
            
        safe_col = make_safe(col)
        
        di = self.header_di()
        
        location = di[safe_col]
        
        already_used = []
        
        indexes = []
        
        for x,r in enumerate(self.data):
            o = r[location]
            h = hash(o)
            if h in already_used:
                continue
            else:
                indexes.append(x)
                already_used.append(h)
        
        return self.__iter__(indexes=indexes)
        
    def ensure_column(self,header_item):
        """
        if column doesn't exist, add it
        """
        if header_item not in self.header:
            self.header.append(header_item)
            return True
        return False

    def xls_book(self):
        """
        create xls book from current ql
        """
        wb = FlexiBook()
        wb.add_sheet_from_ql(self)
        return wb

    def load_from_generator(self):
        for x,r in enumerate(self.generator):
            if x == 0:
                self.header = r
            else:
                self.data.append(r)        

    def yield_from_generator(self):
        for x,r in enumerate(self.generator):
            if x == 0:
                self.header = r
            else:
                yield r              

    def open(self, filename, tab="", header_row=0,force_unicode=False,start=0,limit=None,codec="",**kwargs):
        """
        populate from a file
        """
        
        #correct for rename of argument
        if "unicode" in kwargs:
            force_unicode = kwargs["unicode"]
        
        if isinstance(filename,list):
            self.filename = os.path.join(*filename)
        else:
            self.filename = filename

        if self.name == "":
            self.name = os.path.splitext(os.path.split(self.filename)[1])[0]   
     
        print("Opening : {0}".format(self.filename))
        ext = os.path.splitext(self.filename)[1]
        if ext == ".xlsx":
            self.generator = import_xlsx(self.filename, tab, header_row)
        elif ext == ".csv":
            self.generator = import_csv(self.filename,force_unicode,start,limit,codec)
        elif ext == ".xls":
            self.generator = import_xls(self.filename, tab, header_row)
            
        if self.cached:
            self.load_from_generator()
        
        return self

    def save(self, filename=None, force_unicode=False):
        """
        save out as a csv or xls
        """
        if filename:
            if isinstance(filename,list):
                file_to_use = os.path.join(*filename)
            else:
                file_to_use = filename
        else:
            file_to_use = self.filename
        
        print("Saving : {0}".format(file_to_use))
        if ".csv" in file_to_use:
            export_csv(file_to_use, self.header, self.data, force_unicode=force_unicode)
        if ".psv" in file_to_use:
            export_csv(file_to_use, self.header, self.data, force_unicode=force_unicode)
        if ".xls" in file_to_use:
            self.xls_book().save(file_to_use)
       
    def get_column(self,column_id, unique=False):
        
        previous = []
        
        for r in self:
            v = r[column_id]
            if unique:
                if v not in previous:
                    yield v
                    previous.append(v)
            else:
                yield v

    def use_query(self,query):
        """
        use the header to populate with information out of a django query
        """
        self.data = [[y for y in x] for x in query.values_list(*self.header)]
        return self

    def header_di(self):
        """
        returns a dictionary of 'safe' header (striped, lowered) and positions. 
        """
        def make_safe(v):
            if isinstance(v,float):
                return str(v).lower().strip()
            if v:
                return v.lower().strip()
            else:
                return ""

        return {make_safe(h): x for x, h in enumerate(self.header)}

    def remap(self,new_header):
        """
        return a new quickgrid with a reduced or changed header
        """
        new = self.__class__()
        new.header = new_header
        for r in self:
            new.add([r[x] for x in new.header])
        return new
            
    def __getitem__(self,key):
            header_dict = self.header_di()
            return ItemRow(self,header_dict,self.data[key])     

    def __len__(self):
        return len(self.data)

    def iterator_source(self):
        if self.cached:
            return self.data
        else:
            return self.yield_from_generator()

    def __iter__(self,indexes=None):
        """
        generator that returns an object that's
         __getitem__ accepts the column name to return a cell instance.
        
        so you can go:
        
        for r in ql:
            p = r["person"]
            
        to get the info out of the person column
         
        """
        header_dict = None

        # puts back any changes into the main generator to update the
        # QuickList 
        for x,r in enumerate(self.iterator_source()):
            if header_dict == None:
                header_dict = self.header_di() 
            if indexes and x not in indexes: # if only yielding certain things, ignore all others
                continue
            ir = ItemRow(self,header_dict,r)
            for col, func in self.transformers:
                ir[col] = func(ir[col])
            yield ir
            r[:] = ir[:]
            
